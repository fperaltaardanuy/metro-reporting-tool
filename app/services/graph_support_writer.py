from __future__ import annotations

import calendar
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.db.models import (
    ApprovalStatus,
    FunctionalArea,
    MonthlyBudget,
    PlanningLine,
    PlanningTimeValue,
    Request,
    RequestStatus,
    ReportCode,
    ServiceActivityType,
)
from app.services.monthly_template_writer import (
    FIRST_MONTH_COLUMN,
    INDICATOR_ID_COLUMN,
    MONTH_HEADER_ROW,
    MONTHLY_SHEET_NAME,
)

SUPPORT_SHEET_NAME = "Soporte-graficas"
LABEL_COLUMN = 1
FIRST_DATA_COLUMN = 2


@dataclass(frozen=True)
class GraphSupportWriteResult:
    sheet_name: str
    graph_count: int
    last_row: int


@dataclass(frozen=True)
class SeriesDefinition:
    label: str
    indicator_id: str | None = None
    source_row: int | None = None
    cumulative: bool = False
    percentage: bool = False


class GraphSupportWriter:
    MONTH_LABELS = [
        "ene",
        "feb",
        "mar",
        "abr",
        "may",
        "jun",
        "jul",
        "ago",
        "sep",
        "oct",
        "nov",
        "dic",
    ]

    GENERAL_AREA_NAMES = ["SISIE", "ÁREA", "CISIC", "OGP", "SISF"]

    def write_support_sheet(
        self,
        workbook_path: str | Path,
        session: Session,
        year: int,
        end_month: int,
    ) -> GraphSupportWriteResult:
        workbook_path = Path(workbook_path)
        workbook = load_workbook(workbook_path)
        source_sheet = self._get_source_sheet(workbook)

        if SUPPORT_SHEET_NAME in workbook.sheetnames:
            del workbook[SUPPORT_SHEET_NAME]

        sheet = workbook.create_sheet(SUPPORT_SHEET_NAME)
        self._configure_sheet(sheet)

        workbook_month_columns = self._get_workbook_month_columns(source_sheet, year)
        if not any(month <= end_month for month in workbook_month_columns):
            raise ValueError(
                "No se han encontrado columnas mensuales del año seleccionado en el CM."
            )
        month_columns = self._build_support_month_columns(
            workbook_month_columns=workbook_month_columns,
            year=year,
            end_month=end_month,
        )

        current_row = 1
        graph_count = 0
        end_date = self._month_end(year, end_month)

        current_row = self._write_monthly_graph(
            sheet,
            current_row,
            title=f"ST Creadas en {year}",
            months=[month for _, month, _ in month_columns],
            series=[
                SeriesDefinition(
                    label="IN20-EFIC-II",
                    indicator_id="IN20-EFIC-II",
                    cumulative=True,
                )
            ],
            source_sheet=source_sheet,
            month_columns=month_columns,
        )
        graph_count += 1

        first_monthly_graphs = [
            (
                "ST entregadas y % ST entregadas",
                [
                    SeriesDefinition(label="IN22-EFIC-IA", indicator_id="IN22-EFIC-IA"),
                    SeriesDefinition(
                        label="IN27-EFIC-II",
                        indicator_id="IN27-EFIC-II",
                        percentage=True,
                    ),
                ],
            ),
            (
                "ST en curso",
                [SeriesDefinition(label="IN21-EFIC-IA", indicator_id="IN21-EFIC-IA")],
            ),
            (
                "ST cerradas",
                [
                    SeriesDefinition(
                        label="IN23-EFIC-IA",
                        indicator_id="IN23-EFIC-IA",
                        cumulative=True,
                    )
                ],
            ),
        ]

        for title, series in first_monthly_graphs:
            current_row = self._write_monthly_graph(
                sheet,
                current_row,
                title=title,
                months=[month for _, month, _ in month_columns],
                series=series,
                source_sheet=source_sheet,
                month_columns=month_columns,
            )
            graph_count += 1

        first_sqlite_graphs = [
            (
                "Tipo de actividad del servicio",
                self._get_service_activity_percentages(session, year, end_date),
                "Tipo de actividad",
                "Porcentaje de ST",
                True,
            ),
            (
                "Grupo de interés",
                self._get_interest_group_percentages(session, year, end_date),
                "Grupo de interés",
                "Porcentaje de ST",
                True,
            ),
            (
                "Ámbito funcional",
                self._get_functional_area_counts(session, year, end_date),
                "Área funcional",
                "Número de ST",
                False,
            ),
            (
                "Sistemas",
                self._get_system_counts(session, year, end_date),
                "Sistema",
                "Número de ST",
                False,
            ),
        ]

        for title, values, header, value_label, percentage in first_sqlite_graphs:
            current_row = self._write_category_graph(
                sheet,
                current_row,
                title=title,
                categories=[label for label, _ in values],
                values=[value for _, value in values],
                header_label=header,
                value_label=value_label,
                percentage=percentage,
            )
            graph_count += 1

        middle_monthly_graphs = [
            (
                "Solicitudes de cambios de ST",
                [SeriesDefinition(label="IN24-EFIC-IA", indicator_id="IN24-EFIC-IA")],
            ),
            (
                "%Modificaciones detonan nueva ST",
                [
                    SeriesDefinition(
                        label="IN08-EFEC-IP",
                        indicator_id="IN08-EFEC-IP",
                        percentage=True,
                    )
                ],
            ),
            (
                "%ST Modificadas respecto a aprobadas",
                [
                    SeriesDefinition(
                        label="IN07-EFEC-IP",
                        indicator_id="IN07-EFEC-IP",
                        percentage=True,
                    )
                ],
            ),
        ]

        for title, series in middle_monthly_graphs:
            current_row = self._write_monthly_graph(
                sheet,
                current_row,
                title=title,
                months=[month for _, month, _ in month_columns],
                series=series,
                source_sheet=source_sheet,
                month_columns=month_columns,
            )
            graph_count += 1

        current_row = self._write_category_graph(
            sheet,
            current_row,
            title="Totales de ST para porcentajes",
            categories=["Total de ST", "ST en curso", "ST cerradas", "ST anuladas"],
            values=[
                self._get_total_request_count(session),
                self._get_open_request_count(session),
                self._get_status_request_count(session, "Cerrada"),
                self._get_status_request_count(session, "Cancelada"),
            ],
            header_label="Tipo",
            value_label="Número de ST",
            percentage=False,
        )
        graph_count += 1

        area_sqlite_graphs = [
            (
                "Total de ST por área",
                self._get_general_area_percentages(session, None),
                "Área funcional",
                "Porcentaje sobre total",
                True,
            ),
            (
                "ST en curso por área",
                self._get_general_area_percentages(session, "open"),
                "Área funcional",
                "Porcentaje sobre total",
                True,
            ),
            (
                "ST cerradas por área",
                self._get_general_area_percentages(session, "closed"),
                "Área funcional",
                "Porcentaje sobre total",
                True,
            ),
            (
                "ST anuladas por área",
                self._get_general_area_percentages(session, "cancelled"),
                "Área funcional",
                "Porcentaje sobre total",
                True,
            ),
        ]

        for title, values, header, value_label, percentage in area_sqlite_graphs:
            current_row = self._write_category_graph(
                sheet,
                current_row,
                title=title,
                categories=[label for label, _ in values],
                values=[value for _, value in values],
                header_label=header,
                value_label=value_label,
                percentage=percentage,
            )
            graph_count += 1

        final_monthly_graphs = [
            (
                "Cumplimiento de la Planificación estratégica",
                [
                    SeriesDefinition(
                        label="IN02-EFEC-IL",
                        indicator_id="IN02-EFEC-IL",
                        percentage=True,
                    )
                ],
            ),
            (
                "Cumplimiento de la Planificación funcional",
                [
                    SeriesDefinition(
                        label="IDATGENAGD01",
                        source_row=12,
                        percentage=True,
                    ),
                    SeriesDefinition(
                        label="IDATGENAGN02",
                        source_row=13,
                        percentage=True,
                    ),
                    SeriesDefinition(
                        label="IDATGENPRE01",
                        source_row=14,
                        percentage=True,
                    ),
                    SeriesDefinition(
                        label="IDATGENSSP01",
                        source_row=15,
                        percentage=True,
                    ),
                    SeriesDefinition(
                        label="IDATGENDEL02",
                        source_row=17,
                        percentage=True,
                    ),
                ],
            ),
            (
                "%ST Entregadas con desviación de plazo",
                [
                    SeriesDefinition(
                        label="IN05-EFEC-IL",
                        indicator_id="IN05-EFEC-IL",
                        percentage=True,
                    )
                ],
            ),
            (
                "%ST Entregadas con desviación de presupuesto",
                [
                    SeriesDefinition(
                        label="IN06-EFEC-IL",
                        indicator_id="IN06-EFEC-IL",
                        percentage=True,
                    )
                ],
            ),
            (
                "Tasa de media de desviación de plazo",
                [
                    SeriesDefinition(
                        label="IN12-EFEC-IL",
                        indicator_id="IN12-EFEC-IL",
                        percentage=True,
                    )
                ],
            ),
            (
                "Tasa de media de desviación de presupuesto",
                [
                    SeriesDefinition(
                        label="IN10-EFEC-IL",
                        indicator_id="IN10-EFEC-IL",
                        percentage=True,
                    )
                ],
            ),
        ]

        for title, series in final_monthly_graphs:
            current_row = self._write_monthly_graph(
                sheet,
                current_row,
                title=title,
                months=[month for _, month, _ in month_columns],
                series=series,
                source_sheet=source_sheet,
                month_columns=month_columns,
            )
            graph_count += 1

        current_row = self._write_budget_planning_graph(
            sheet,
            current_row,
            months=[month for _, month, _ in month_columns],
            estimated_values=self._get_planning_cost_by_month(
                session, year, end_month, "estimated"
            ),
            budget_values=self._get_monthly_budget_by_month(session, year, end_month),
            real_values=self._get_planning_cost_by_month(
                session, year, end_month, "real"
            ),
        )
        graph_count += 1

        current_row = self._write_monthly_graph(
            sheet,
            current_row,
            title="Cumplimiento Planificación Presupuestaria",
            months=[month for _, month, _ in month_columns],
            series=[
                SeriesDefinition(
                    label="IN01-EFEC-IL",
                    indicator_id="IN01-EFEC-IL",
                    percentage=True,
                )
            ],
            source_sheet=source_sheet,
            month_columns=month_columns,
        )
        graph_count += 1

        current_row = self._write_billing_graph(
            sheet,
            current_row,
            year=year,
            end_month=end_month,
            months=[month for _, month, _ in month_columns],
            monthly_values=self._get_planning_cost_by_month(
                session, year, end_month, "real"
            ),
            annual_target=self._get_annual_budget(session, year),
        )
        graph_count += 1

        final_service_monthly_graphs = [
            (
                "Número de ETCs en ejecución",
                [
                    SeriesDefinition(
                        label="IN19-CALS-IA",
                        indicator_id="IN19-CALS-IA",
                    )
                ],
            ),
            (
                "Dedicación del Director de Servicio",
                [
                    SeriesDefinition(
                        label="IN28-EFIC-IA",
                        indicator_id="IN28-EFIC-IA",
                        percentage=True,
                    )
                ],
            ),
            (
                "Tiempo de respuesta en la valoración de solicitudes <= 48h",
                [
                    SeriesDefinition(
                        label="IN04-EFEC-IP",
                        indicator_id="IN04-EFEC-IP",
                        percentage=True,
                    )
                ],
            ),
        ]

        for title, series in final_service_monthly_graphs:
            current_row = self._write_monthly_graph(
                sheet,
                current_row,
                title=title,
                months=[month for _, month, _ in month_columns],
                series=series,
                source_sheet=source_sheet,
                month_columns=month_columns,
            )
            graph_count += 1

        workbook.save(workbook_path)
        return GraphSupportWriteResult(
            sheet_name=SUPPORT_SHEET_NAME,
            graph_count=graph_count,
            last_row=current_row - 1,
        )

    def _write_monthly_graph(
        self,
        sheet,
        start_row: int,
        title: str,
        months: list[date],
        series: list[SeriesDefinition],
        source_sheet,
        month_columns: list[tuple[int | None, date, bool]],
    ) -> int:
        title_row = start_row
        header_row = start_row + 1

        end_column = max(LABEL_COLUMN, FIRST_DATA_COLUMN + len(months) - 1)
        self._write_title(sheet, title_row, title, end_column=end_column)
        sheet.cell(row=header_row, column=LABEL_COLUMN).value = "Mes"
        self._style_header_cell(sheet.cell(row=header_row, column=LABEL_COLUMN))

        for index, month in enumerate(months, start=FIRST_DATA_COLUMN):
            cell = sheet.cell(row=header_row, column=index)
            cell.value = self._format_month(month)
            self._style_header_cell(cell)

        indicator_row_map = self._build_indicator_row_map(source_sheet)
        row = header_row + 1

        for definition in series:
            sheet.cell(row=row, column=LABEL_COLUMN).value = definition.label
            self._style_label_cell(sheet.cell(row=row, column=LABEL_COLUMN))

            source_row = definition.source_row
            if source_row is None and definition.indicator_id is not None:
                source_row = indicator_row_map.get(definition.indicator_id)

            values = self._read_values_from_row(source_sheet, source_row, month_columns)
            if definition.cumulative:
                values = self._accumulate(
                    values=values,
                    force_zero_flags=[force_zero for _, _, force_zero in month_columns],
                )

            for index, value in enumerate(values, start=FIRST_DATA_COLUMN):
                cell = sheet.cell(row=row, column=index)
                cell.value = value
                if definition.percentage:
                    cell.number_format = "0%"

            row += 1

        return row + 2

    def _write_budget_planning_graph(
        self,
        sheet,
        start_row: int,
        months: list[date],
        estimated_values: list[float],
        budget_values: list[float],
        real_values: list[float],
    ) -> int:
        title_row = start_row
        header_row = start_row + 1
        first_values_row = start_row + 2

        end_column = max(LABEL_COLUMN, FIRST_DATA_COLUMN + len(months) - 1)
        self._write_title(
            sheet,
            title_row,
            "Planificación Presupuestaria",
            end_column=end_column,
        )
        sheet.cell(row=header_row, column=LABEL_COLUMN).value = "Mes"
        self._style_header_cell(sheet.cell(row=header_row, column=LABEL_COLUMN))

        for index, month in enumerate(months, start=FIRST_DATA_COLUMN):
            cell = sheet.cell(row=header_row, column=index)
            cell.value = self._format_month(month)
            self._style_header_cell(cell)

        rows = [
            ("Planificación Estratégica", estimated_values),
            ("Planificación Económica", budget_values),
            ("Ejecución Material", real_values),
        ]

        for row_offset, (label, values) in enumerate(rows):
            row = first_values_row + row_offset
            sheet.cell(row=row, column=LABEL_COLUMN).value = label
            self._style_label_cell(sheet.cell(row=row, column=LABEL_COLUMN))

            for column_index, value in enumerate(values, start=FIRST_DATA_COLUMN):
                cell = sheet.cell(row=row, column=column_index)
                cell.value = value
                cell.number_format = '#,##0.00 "€"'

        return first_values_row + len(rows) + 2

    def _write_billing_graph(
        self,
        sheet,
        start_row: int,
        year: int,
        end_month: int,
        months: list[date],
        monthly_values: list[float],
        annual_target: float,
    ) -> int:
        title_row = start_row
        header_row = start_row + 1
        accumulated_row = start_row + 2
        monthly_row = start_row + 3
        target_row = start_row + 4

        end_column = max(LABEL_COLUMN, FIRST_DATA_COLUMN + len(months) - 1)
        self._write_title(
            sheet, title_row, f"Facturación {year}", end_column=end_column
        )
        sheet.cell(row=header_row, column=LABEL_COLUMN).value = "Mes"
        self._style_header_cell(sheet.cell(row=header_row, column=LABEL_COLUMN))

        for index, month in enumerate(months, start=FIRST_DATA_COLUMN):
            cell = sheet.cell(row=header_row, column=index)
            cell.value = self._format_month(month)
            self._style_header_cell(cell)

        accumulated_values = self._accumulate(
            values=monthly_values,
            force_zero_flags=[month.month > end_month for month in months],
        )
        rows = [
            ("Facturación Acumulada", accumulated_values),
            ("Facturación Mensual", monthly_values),
            (f"Facturación objetivo {year}", [annual_target for _ in months]),
        ]

        for row_index, (label, values) in zip(
            [accumulated_row, monthly_row, target_row],
            rows,
        ):
            sheet.cell(row=row_index, column=LABEL_COLUMN).value = label
            self._style_label_cell(sheet.cell(row=row_index, column=LABEL_COLUMN))

            for column_index, value in enumerate(values, start=FIRST_DATA_COLUMN):
                cell = sheet.cell(row=row_index, column=column_index)
                cell.value = value
                cell.number_format = '#,##0.00 "€"'

        return target_row + 3

    def _write_category_graph(
        self,
        sheet,
        start_row: int,
        title: str,
        categories: list[str],
        values: list[float],
        header_label: str,
        value_label: str,
        percentage: bool,
    ) -> int:
        title_row = start_row
        header_row = start_row + 1
        values_row = start_row + 2

        end_column = max(LABEL_COLUMN, FIRST_DATA_COLUMN + len(categories) - 1)
        self._write_title(sheet, title_row, title, end_column=end_column)
        sheet.cell(row=header_row, column=LABEL_COLUMN).value = header_label
        sheet.cell(row=values_row, column=LABEL_COLUMN).value = value_label
        self._style_header_cell(sheet.cell(row=header_row, column=LABEL_COLUMN))
        self._style_label_cell(sheet.cell(row=values_row, column=LABEL_COLUMN))

        for index, category in enumerate(categories, start=FIRST_DATA_COLUMN):
            cell = sheet.cell(row=header_row, column=index)
            cell.value = category
            self._style_header_cell(cell)

        for index, value in enumerate(values, start=FIRST_DATA_COLUMN):
            cell = sheet.cell(row=values_row, column=index)
            cell.value = value
            if percentage:
                cell.number_format = "0%"

        return values_row + 3

    def _get_source_sheet(self, workbook):
        if MONTHLY_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"No existe la hoja '{MONTHLY_SHEET_NAME}' en el fichero seleccionado."
            )
        return workbook[MONTHLY_SHEET_NAME]

    def _configure_sheet(self, sheet) -> None:
        sheet.freeze_panes = "B2"
        sheet.column_dimensions["A"].width = 38
        for column in range(FIRST_DATA_COLUMN, 26):
            sheet.column_dimensions[get_column_letter(column)].width = 14

    def _write_title(self, sheet, row: int, title: str, end_column: int) -> None:
        cell = sheet.cell(row=row, column=LABEL_COLUMN)
        cell.value = title
        cell.font = Font(bold=True, size=13, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="left")

        if end_column > LABEL_COLUMN:
            sheet.merge_cells(
                start_row=row,
                start_column=LABEL_COLUMN,
                end_row=row,
                end_column=end_column,
            )

    def _style_header_cell(self, cell) -> None:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    def _style_label_cell(self, cell) -> None:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    def _get_workbook_month_columns(
        self,
        sheet,
        year: int,
    ) -> dict[int, int]:
        result: dict[int, int] = {}

        for column in range(FIRST_MONTH_COLUMN, sheet.max_column + 1):
            month = self._normalize_month(
                sheet.cell(row=MONTH_HEADER_ROW, column=column).value
            )
            if month is None:
                continue

            if month.year == year:
                result[month.month] = column

        return result

    def _build_support_month_columns(
        self,
        workbook_month_columns: dict[int, int],
        year: int,
        end_month: int,
    ) -> list[tuple[int | None, date, bool]]:
        return [
            (
                workbook_month_columns.get(month) if month <= end_month else None,
                date(year, month, 1),
                month > end_month,
            )
            for month in range(1, 13)
        ]

    def _normalize_month(self, value: Any) -> date | None:
        if isinstance(value, datetime):
            return date(value.year, value.month, 1)

        if isinstance(value, date):
            return date(value.year, value.month, 1)

        return None

    def _build_indicator_row_map(self, sheet) -> dict[str, int]:
        aliases = {
            "IN20-EFIC-IA": "IN20-EFIC-II",
        }
        result: dict[str, int] = {}

        for row in range(1, sheet.max_row + 1):
            raw_value = sheet.cell(row=row, column=INDICATOR_ID_COLUMN).value
            if raw_value is None:
                continue

            indicator_id = str(raw_value).strip()
            if indicator_id:
                result[indicator_id] = row

        for alias, target in aliases.items():
            if target in result:
                result[alias] = result[target]

        return result

    def _read_values_from_row(
        self,
        sheet,
        row: int | None,
        month_columns: list[tuple[int | None, date, bool]],
    ) -> list[float | None]:
        if row is None:
            return [None for _ in month_columns]

        return [
            (
                0.0
                if force_zero or column is None
                else self._normalize_cell_value(
                    sheet.cell(row=row, column=column).value
                )
            )
            for column, _, force_zero in month_columns
        ]

    def _normalize_cell_value(self, value: Any) -> float | None:
        if isinstance(value, bool):
            return None

        if isinstance(value, int | float):
            return float(value)

        if isinstance(value, str):
            normalized = value.strip().lower()
            if normalized in {"", "-", "error"}:
                return None

        return None

    def _accumulate(
        self,
        values: list[float | None],
        force_zero_flags: list[bool],
    ) -> list[float | None]:
        result: list[float | None] = []
        total = 0.0

        for value, force_zero in zip(values, force_zero_flags):
            if force_zero:
                result.append(0.0)
                continue

            if value is not None:
                total += value
            result.append(total)

        return result

    def _format_month(self, month: date) -> str:
        return f"{self.MONTH_LABELS[month.month - 1]}-{str(month.year)[2:]}"

    def _month_end(self, year: int, month: int) -> date:
        return date(year, month, calendar.monthrange(year, month)[1])

    def _requests_created_ytd_query(self, session: Session, year: int, end_date: date):
        return (
            session.query(Request)
            .filter(Request.request_date.isnot(None))
            .filter(Request.request_date >= date(year, 1, 1))
            .filter(Request.request_date <= end_date)
        )

    def _get_service_activity_percentages(
        self, session: Session, year: int, end_date: date
    ) -> list[tuple[str, float]]:
        rows = (
            self._requests_created_ytd_query(session, year, end_date)
            .outerjoin(
                ServiceActivityType,
                Request.service_activity_type_id == ServiceActivityType.id,
            )
            .with_entities(
                ServiceActivityType.name,
                func.count(Request.id),
            )
            .group_by(ServiceActivityType.name)
            .order_by(func.count(Request.id).desc(), ServiceActivityType.name)
            .all()
        )
        return self._to_percentages(
            {name or "Sin tipo de actividad": int(count or 0) for name, count in rows}
        )

    def _get_functional_area_counts(
        self, session: Session, year: int, end_date: date
    ) -> list[tuple[str, float]]:
        rows = (
            self._requests_created_ytd_query(session, year, end_date)
            .outerjoin(FunctionalArea, Request.functional_area_id == FunctionalArea.id)
            .with_entities(FunctionalArea.name, func.count(Request.id))
            .group_by(FunctionalArea.name)
            .order_by(func.count(Request.id).desc(), FunctionalArea.name)
            .all()
        )
        return [
            (name or "Sin área funcional", float(count or 0)) for name, count in rows
        ]

    def _get_interest_group_percentages(
        self, session: Session, year: int, end_date: date
    ) -> list[tuple[str, float]]:
        requests = self._requests_created_ytd_query(session, year, end_date).all()
        combinations: dict[str, int] = {}

        for request in requests:
            names = sorted(
                group.name
                for group in request.interest_group_activity_types
                if group.name
            )
            label = "; ".join(names) if names else "Sin grupo de interés"
            combinations[label] = combinations.get(label, 0) + 1

        return self._to_percentages(combinations)

    def _get_system_counts(
        self, session: Session, year: int, end_date: date
    ) -> list[tuple[str, float]]:
        requests = self._requests_created_ytd_query(session, year, end_date).all()
        combinations: dict[str, int] = {}

        for request in requests:
            names = sorted(system.name for system in request.systems if system.name)
            label = "; ".join(names) if names else "Sin sistema"
            combinations[label] = combinations.get(label, 0) + 1

        return [
            (label, float(count))
            for label, count in sorted(
                combinations.items(),
                key=lambda item: (-item[1], item[0]),
            )
        ]

    def _get_general_area_percentages(
        self,
        session: Session,
        status_filter: str | None,
    ) -> list[tuple[str, float]]:
        query = session.query(Request).outerjoin(
            FunctionalArea,
            Request.functional_area_id == FunctionalArea.id,
        )

        if status_filter == "open":
            query = (
                query.join(
                    ApprovalStatus,
                    Request.approval_status_id == ApprovalStatus.id,
                )
                .outerjoin(
                    RequestStatus,
                    Request.request_status_id == RequestStatus.id,
                )
                .filter(ApprovalStatus.name == "MdM Aprobada")
                .filter(
                    or_(
                        RequestStatus.name.is_(None),
                        ~RequestStatus.name.in_(["Cancelada", "Cerrada", "Rechazada"]),
                    )
                )
            )
        elif status_filter == "closed":
            query = query.join(
                RequestStatus,
                Request.request_status_id == RequestStatus.id,
            ).filter(RequestStatus.name == "Cerrada")
        elif status_filter == "cancelled":
            query = query.join(
                RequestStatus,
                Request.request_status_id == RequestStatus.id,
            ).filter(RequestStatus.name == "Cancelada")

        counts = {area: 0 for area in self.GENERAL_AREA_NAMES}
        counts["OTRAS"] = 0

        for request in query.all():
            area_name = request.functional_area.name if request.functional_area else ""
            general_area = self._general_area_name(area_name)
            counts[general_area] = counts.get(general_area, 0) + 1

        return self._to_percentages(counts)

    def _get_total_request_count(self, session: Session) -> float:
        result = session.query(func.count(Request.id)).scalar()
        return float(result or 0)

    def _get_open_request_count(self, session: Session) -> float:
        result = (
            session.query(func.count(Request.id))
            .join(
                ApprovalStatus,
                Request.approval_status_id == ApprovalStatus.id,
            )
            .outerjoin(
                RequestStatus,
                Request.request_status_id == RequestStatus.id,
            )
            .filter(ApprovalStatus.name == "MdM Aprobada")
            .filter(
                or_(
                    RequestStatus.name.is_(None),
                    ~RequestStatus.name.in_(["Cancelada", "Cerrada", "Rechazada"]),
                )
            )
            .scalar()
        )
        return float(result or 0)

    def _get_status_request_count(self, session: Session, status_name: str) -> float:
        result = (
            session.query(func.count(Request.id))
            .join(
                RequestStatus,
                Request.request_status_id == RequestStatus.id,
            )
            .filter(RequestStatus.name == status_name)
            .scalar()
        )
        return float(result or 0)

    def _general_area_name(self, area_name: str) -> str:
        normalized = area_name.strip().upper()

        if normalized.startswith("ÁREA") or normalized.startswith("AREA"):
            return "ÁREA"

        for area in ("SISIE", "CISIC", "OGP", "SISF"):
            if normalized.startswith(area):
                return area

        return "OTRAS"

    def _to_percentages(self, counts: dict[str, int]) -> list[tuple[str, float]]:
        total = sum(counts.values())
        if total == 0:
            return [(label, 0.0) for label in counts]

        return [
            (label, count / total)
            for label, count in sorted(
                counts.items(),
                key=lambda item: (-item[1], item[0]),
            )
            if count > 0
        ]

    def _get_planning_cost_by_month(
        self,
        session: Session,
        year: int,
        end_month: int,
        source_type: str,
    ) -> list[float]:
        values = [0.0 for _ in range(12)]
        rows = (
            session.query(
                PlanningTimeValue.month,
                PlanningLine.report_code,
                func.sum(PlanningTimeValue.hours),
                ReportCode.unit_price,
                ReportCode.at_unit_hours,
            )
            .select_from(PlanningLine)
            .join(
                PlanningTimeValue,
                PlanningTimeValue.planning_line_id == PlanningLine.id,
            )
            .join(
                ReportCode,
                ReportCode.code == PlanningLine.report_code,
            )
            .filter(PlanningLine.source_type == source_type)
            .filter(PlanningLine.report_code.isnot(None))
            .filter(PlanningLine.report_code != "")
            .filter(PlanningTimeValue.year == year)
            .filter(PlanningTimeValue.month <= end_month)
            .group_by(
                PlanningTimeValue.month,
                PlanningLine.report_code,
                ReportCode.unit_price,
                ReportCode.at_unit_hours,
            )
            .all()
        )

        for month, _report_code, total_hours, unit_price, at_unit_hours in rows:
            if month is None or month < 1 or month > 12:
                continue

            hours = float(total_hours or 0.0)
            if hours == 0 or unit_price is None or at_unit_hours is None:
                continue

            at_unit_hours_value = float(at_unit_hours)
            if at_unit_hours_value == 0:
                continue

            values[month - 1] += hours * (float(unit_price) / at_unit_hours_value)

        return values

    def _get_monthly_budget_by_month(
        self,
        session: Session,
        year: int,
        end_month: int,
    ) -> list[float]:
        values = [0.0 for _ in range(12)]
        rows = (
            session.query(MonthlyBudget.month, MonthlyBudget.amount)
            .filter(MonthlyBudget.year == year)
            .filter(MonthlyBudget.month <= end_month)
            .filter(MonthlyBudget.amount.isnot(None))
            .all()
        )

        for month, amount in rows:
            if month is None or month < 1 or month > 12:
                continue

            values[month - 1] = float(amount or 0.0)

        return values

    def _get_annual_budget(self, session: Session, year: int) -> float:
        value = (
            session.query(func.sum(MonthlyBudget.amount))
            .filter(MonthlyBudget.year == year)
            .filter(MonthlyBudget.amount.isnot(None))
            .scalar()
        )

        return float(value or 0.0)
