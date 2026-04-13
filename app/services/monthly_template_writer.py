from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.monthly_indicator_service import ReportCodeComplianceValue


MONTHLY_SHEET_NAME = "Cuadro de Mando Principal"

YEAR_HEADER_ROW = 7
MONTH_HEADER_ROW = 8
FIRST_MONTH_COLUMN = 11  # K

INDICATOR_ID_COLUMN = 4  # D
REPORT_CODE_COLUMN = 7  # G

FIRST_INDICATOR_ROW = 9
LAST_INDICATOR_ROW = 43

IN03_FIRST_ROW = 12
IN03_LAST_ROW = 18

PERCENTAGE_INDICATOR_IDS = {
    "IN28-EFIC-IA",
}

@dataclass(frozen=True)
class MonthlyTemplateWriteResult:
    sheet_name: str
    month_column: int
    month_label: str
    written_indicator_ids: list[str]
    written_report_codes: list[str]
    missing_indicator_ids: list[str]
    missing_report_codes: list[str]


class MonthlyTemplateWriter:
    def get_in03_report_codes(self, workbook_path: str | Path) -> list[str]:
        workbook = load_workbook(workbook_path, data_only=True)
        sheet = self._get_monthly_sheet(workbook)
        row_map = self._build_in03_report_code_row_map(sheet)
        return list(row_map.keys())

    def write_monthly_report(
        self,
        workbook_path: str | Path,
        report_month: date,
        indicator_values: dict[str, Any],
        in03_values: list[ReportCodeComplianceValue],
    ) -> MonthlyTemplateWriteResult:
        workbook_path = Path(workbook_path)

        workbook = load_workbook(workbook_path)
        lookup_workbook = load_workbook(workbook_path, data_only=True)

        sheet = self._get_monthly_sheet(workbook)
        lookup_sheet = self._get_monthly_sheet(lookup_workbook)

        month_column = self._get_or_create_month_column(sheet, report_month)
        indicator_row_map = self._build_indicator_row_map(lookup_sheet)
        report_code_row_map = self._build_in03_report_code_row_map(lookup_sheet)

        written_indicator_ids: list[str] = []
        missing_indicator_ids: list[str] = []
        written_report_codes: list[str] = []
        missing_report_codes: list[str] = []

        for indicator_id, value in indicator_values.items():
            self._write_single_indicator(
                sheet=sheet,
                indicator_row_map=indicator_row_map,
                indicator_id=indicator_id,
                value=value,
                month_column=month_column,
                written_indicator_ids=written_indicator_ids,
                missing_indicator_ids=missing_indicator_ids,
            )

        in03_by_report_code = {
            item.report_code.strip(): item
            for item in in03_values
            if item.report_code is not None and item.report_code.strip()
        }

        for report_code, row in report_code_row_map.items():
            item = in03_by_report_code.get(report_code)
            if item is None:
                missing_report_codes.append(report_code)
                continue

            excel_value = self._normalize_in03_value_for_excel(item.value)
            sheet.cell(row=row, column=month_column).value = excel_value
            written_report_codes.append(report_code)

        workbook.save(workbook_path)

        return MonthlyTemplateWriteResult(
            sheet_name=sheet.title,
            month_column=month_column,
            month_label=report_month.strftime("%m-%Y"),
            written_indicator_ids=written_indicator_ids,
            written_report_codes=written_report_codes,
            missing_indicator_ids=missing_indicator_ids,
            missing_report_codes=missing_report_codes,
        )

    def _get_monthly_sheet(self, workbook) -> Worksheet:
        if MONTHLY_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"No existe la hoja '{MONTHLY_SHEET_NAME}' en el fichero seleccionado."
            )

        return workbook[MONTHLY_SHEET_NAME]

    def _build_indicator_row_map(self, sheet: Worksheet) -> dict[str, int]:
        result: dict[str, int] = {}

        for row in range(FIRST_INDICATOR_ROW, LAST_INDICATOR_ROW + 1):
            value = sheet.cell(row=row, column=INDICATOR_ID_COLUMN).value
            if value is None:
                continue

            indicator_id = str(value).strip()
            if not indicator_id.startswith("IN"):
                continue

            result[indicator_id] = row

        if not result:
            raise ValueError(
                "No se han encontrado IDs de indicadores en la columna D."
            )

        return result

    def _build_in03_report_code_row_map(self, sheet: Worksheet) -> dict[str, int]:
        result: dict[str, int] = {}

        for row in range(IN03_FIRST_ROW, IN03_LAST_ROW + 1):
            value = sheet.cell(row=row, column=REPORT_CODE_COLUMN).value
            if value is None:
                continue

            report_code = str(value).strip()
            if not report_code:
                continue

            result[report_code] = row

        if not result:
            raise ValueError(
                "No se han encontrado report codes del bloque IN03 en la columna G."
            )

        return result

    def _write_single_indicator(
        self,
        sheet: Worksheet,
        indicator_row_map: dict[str, int],
        indicator_id: str,
        value: Any,
        month_column: int,
        written_indicator_ids: list[str],
        missing_indicator_ids: list[str],
    ) -> None:
        row = indicator_row_map.get(indicator_id)
        if row is None:
            missing_indicator_ids.append(indicator_id)
            return

        excel_value = self._normalize_indicator_value_for_excel(indicator_id, value)
        sheet.cell(row=row, column=month_column).value = excel_value
        written_indicator_ids.append(indicator_id)
    
    def _normalize_indicator_value_for_excel(
        self,
        indicator_id: str,
        value: Any,
    ) -> Any:
        if isinstance(value, str):
            return value

        if indicator_id in PERCENTAGE_INDICATOR_IDS:
            return value / 100.0

        return value

    def _normalize_in03_value_for_excel(self, value: float | str) -> float | str:
        if isinstance(value, str):
            return value

        # El servicio devuelve porcentaje 0..100, pero Excel tiene formato 0%
        # por lo que en la celda hay que escribir 0..1.
        return value / 100.0

    def _get_or_create_month_column(self, sheet: Worksheet, report_month: date) -> int:
        existing_column = self._find_existing_month_column(sheet, report_month)
        if existing_column is not None:
            return existing_column

        last_month_column = self._find_last_month_column(sheet)
        last_month_date = self._get_month_date_from_cell(
            sheet.cell(row=MONTH_HEADER_ROW, column=last_month_column).value
        )

        if last_month_date is None:
            raise ValueError(
                "No se ha podido interpretar la última columna mensual de la plantilla."
            )

        expected_next_month = self._add_one_month(last_month_date)
        if (report_month.year, report_month.month) != (
            expected_next_month.year,
            expected_next_month.month,
        ):
            raise ValueError(
                "La plantilla no contiene el mes solicitado y este no es el siguiente mes "
                "al último existente. Por ahora solo se soporta añadir el siguiente mes consecutivo."
            )

        new_column = last_month_column + 1

        self._copy_column_layout(
            sheet=sheet,
            source_column=last_month_column,
            target_column=new_column,
        )

        self._write_month_headers(
            sheet=sheet,
            target_column=new_column,
            report_month=report_month,
            previous_column=last_month_column,
            previous_month=last_month_date,
        )

        return new_column

    def _find_existing_month_column(self, sheet: Worksheet, report_month: date) -> int | None:
        for column in range(FIRST_MONTH_COLUMN, sheet.max_column + 1):
            raw_value = sheet.cell(row=MONTH_HEADER_ROW, column=column).value
            month_date = self._get_month_date_from_cell(raw_value)
            if month_date is None:
                continue

            if month_date.year == report_month.year and month_date.month == report_month.month:
                return column

        return None

    def _find_last_month_column(self, sheet: Worksheet) -> int:
        last_column = FIRST_MONTH_COLUMN - 1

        for column in range(FIRST_MONTH_COLUMN, sheet.max_column + 1):
            raw_value = sheet.cell(row=MONTH_HEADER_ROW, column=column).value
            month_date = self._get_month_date_from_cell(raw_value)
            if month_date is not None:
                last_column = column

        if last_column < FIRST_MONTH_COLUMN:
            raise ValueError(
                "No se han encontrado columnas mensuales en la fila 8 a partir de la columna K."
            )

        return last_column

    def _write_month_headers(
        self,
        sheet: Worksheet,
        target_column: int,
        report_month: date,
        previous_column: int,
        previous_month: date,
    ) -> None:
        sheet.cell(
            row=MONTH_HEADER_ROW,
            column=target_column,
        ).value = datetime(report_month.year, report_month.month, 1)

        if report_month.year == previous_month.year:
            self._extend_year_header_merge(
                sheet=sheet,
                previous_column=previous_column,
                target_column=target_column,
                year=report_month.year,
            )
        else:
            sheet.cell(row=YEAR_HEADER_ROW, column=target_column).value = report_month.year

    def _extend_year_header_merge(
        self,
        sheet: Worksheet,
        previous_column: int,
        target_column: int,
        year: int,
    ) -> None:
        merge_range = self._find_row_merge_containing_column(
            sheet=sheet,
            row=YEAR_HEADER_ROW,
            column=previous_column,
        )

        if merge_range is None:
            start_column = previous_column
            start_coordinate = f"{get_column_letter(start_column)}{YEAR_HEADER_ROW}"
            sheet[start_coordinate] = year
            sheet.merge_cells(
                start_row=YEAR_HEADER_ROW,
                start_column=start_column,
                end_row=YEAR_HEADER_ROW,
                end_column=target_column,
            )
            return

        start_column = merge_range.min_col
        sheet.unmerge_cells(str(merge_range))
        sheet.merge_cells(
            start_row=YEAR_HEADER_ROW,
            start_column=start_column,
            end_row=YEAR_HEADER_ROW,
            end_column=target_column,
        )
        sheet.cell(row=YEAR_HEADER_ROW, column=start_column).value = year

    def _find_row_merge_containing_column(
        self,
        sheet: Worksheet,
        row: int,
        column: int,
    ):
        for merge_range in sheet.merged_cells.ranges:
            if (
                merge_range.min_row == row
                and merge_range.max_row == row
                and merge_range.min_col <= column <= merge_range.max_col
            ):
                return merge_range

        return None

    def _copy_column_layout(
        self,
        sheet: Worksheet,
        source_column: int,
        target_column: int,
    ) -> None:
        source_letter = get_column_letter(source_column)
        target_letter = get_column_letter(target_column)

        source_width = sheet.column_dimensions[source_letter].width
        if source_width is not None:
            sheet.column_dimensions[target_letter].width = source_width

        for row in range(1, sheet.max_row + 1):
            source_cell = self._get_style_source_cell(sheet, row, source_column)
            target_cell = sheet.cell(row=row, column=target_column)

            if source_cell is None:
                continue

            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)

            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)
            target_cell.number_format = source_cell.number_format

    def _get_style_source_cell(
        self,
        sheet: Worksheet,
        row: int,
        column: int,
    ):
        cell = sheet.cell(row=row, column=column)

        if not isinstance(cell, MergedCell):
            return cell

        for merge_range in sheet.merged_cells.ranges:
            if (
                merge_range.min_row <= row <= merge_range.max_row
                and merge_range.min_col <= column <= merge_range.max_col
            ):
                return sheet.cell(row=merge_range.min_row, column=merge_range.min_col)

        return None

    def _get_month_date_from_cell(self, value: Any) -> date | None:
        if isinstance(value, datetime):
            return date(value.year, value.month, 1)

        if isinstance(value, date):
            return date(value.year, value.month, 1)

        return None

    def _add_one_month(self, value: date) -> date:
        if value.month == 12:
            return date(value.year + 1, 1, 1)

        return date(value.year, value.month + 1, 1)